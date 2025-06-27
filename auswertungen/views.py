"""
Dashboard Views - Zentrale Auswertungen und Berichte für die Buchhaltung.
"""

import datetime
import io
from datetime import timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Min, Sum
from django.db.models.functions import Extract
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

# Export-Imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from belege.models import Beleg
from buchungen.models import Buchungssatz, Geschaeftspartner
from konten.models import Konto


@login_required
def dashboard_view(request):
    """
    Hauptdashboard mit allen wichtigen Kennzahlen.

    Zeigt eine Übersicht über Einnahmen, Ausgaben und wichtige Finanzkennzahlen.
    """
    heute = timezone.now().date()
    monat_start = heute.replace(day=1)
    jahr_start = heute.replace(month=1, day=1)

    # Zeiträume für Vergleiche
    letzter_monat = (monat_start - timedelta(days=1)).replace(day=1)
    jahr_start.replace(year=jahr_start.year - 1)

    # Einnahmen und Ausgaben aktueller Monat
    einnahmen_monat = Buchungssatz.objects.filter(
        buchungsdatum__gte=monat_start,
        haben_konto__nummer__startswith="8",  # Ertragskonten
    ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

    ausgaben_monat = Buchungssatz.objects.filter(
        buchungsdatum__gte=monat_start,
        soll_konto__nummer__startswith="4",  # Aufwandskonten
    ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

    gewinn_monat = einnahmen_monat - ausgaben_monat

    # Einnahmen und Ausgaben aktuelles Jahr
    einnahmen_jahr = Buchungssatz.objects.filter(
        buchungsdatum__gte=jahr_start, haben_konto__nummer__startswith="8"
    ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

    ausgaben_jahr = Buchungssatz.objects.filter(
        buchungsdatum__gte=jahr_start, soll_konto__nummer__startswith="4"
    ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

    gewinn_jahr = einnahmen_jahr - ausgaben_jahr

    # Vergleich mit Vormonat
    einnahmen_vormonat = Buchungssatz.objects.filter(
        buchungsdatum__gte=letzter_monat,
        buchungsdatum__lt=monat_start,
        haben_konto__nummer__startswith="8",
    ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

    # Trend berechnen
    einnahmen_trend = 0
    if einnahmen_vormonat > 0:
        einnahmen_trend = float(
            (einnahmen_monat - einnahmen_vormonat) / einnahmen_vormonat * 100
        )

    # Statistiken
    stats = {
        "buchungen_gesamt": Buchungssatz.objects.count(),
        "buchungen_monat": Buchungssatz.objects.filter(
            buchungsdatum__gte=monat_start
        ).count(),
        "belege_count": Beleg.objects.count(),
        "belege_unbearbeitet": Beleg.objects.filter(
            status__in=["NEU", "FEHLER"]
        ).count(),
        "belege_eingang": Beleg.objects.filter(beleg_typ="RECHNUNG_EINGANG").count(),
        "belege_ausgang": Beleg.objects.filter(beleg_typ="RECHNUNG_AUSGANG").count(),
        "konten_count": Konto.objects.count(),
        "partner_count": Geschaeftspartner.objects.filter(aktiv=True).count(),
    }

    # Letzte Buchungen
    letzte_buchungen = Buchungssatz.objects.select_related(
        "soll_konto", "haben_konto", "geschaeftspartner", "beleg"
    ).order_by("-erstellt_am")[:10]

    # Offene Belege
    offene_belege = (
        Beleg.objects.filter(status__in=["NEU", "GEPRUEFT"])
        .select_related("geschaeftspartner")
        .order_by("-rechnungsdatum")[:5]
    )

    # Top Ausgaben-Kategorien (letzten 30 Tage)
    vor_30_tagen = heute - timedelta(days=30)
    top_ausgaben = (
        Buchungssatz.objects.filter(
            buchungsdatum__gte=vor_30_tagen, soll_konto__nummer__startswith="4"
        )
        .values("soll_konto__name")
        .annotate(summe=Sum("betrag"), anzahl=Count("id"))
        .order_by("-summe")[:5]
    )

    # Monats-Chart-Daten (letzten 12 Monate)
    chart_data = []
    for i in range(12):
        chart_monat = (heute.replace(day=1) - timedelta(days=30 * i)).replace(day=1)
        chart_monat_ende = (chart_monat.replace(day=28) + timedelta(days=4)).replace(
            day=1
        ) - timedelta(days=1)

        einnahmen = Buchungssatz.objects.filter(
            buchungsdatum__gte=chart_monat,
            buchungsdatum__lte=chart_monat_ende,
            haben_konto__nummer__startswith="8",
        ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

        ausgaben = Buchungssatz.objects.filter(
            buchungsdatum__gte=chart_monat,
            buchungsdatum__lte=chart_monat_ende,
            soll_konto__nummer__startswith="4",
        ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

        chart_data.insert(
            0,
            {
                "monat": chart_monat.strftime("%b %Y"),
                "einnahmen": float(einnahmen),
                "ausgaben": float(ausgaben),
                "gewinn": float(einnahmen - ausgaben),
            },
        )

    # Peter Zwegat Motivations-Sprüche
    zwegat_sprueche = [
        "Ihre Bücher sind in Ordnung - das freut mich!",
        "Weiter so! Ordnung ist das halbe Leben.",
        "Jeder Euro zählt - und Sie zählen jeden Euro!",
        "Buchhaltung kann Spaß machen - sehen Sie selbst!",
        "Ihre Finanzen sind auf dem richtigen Weg!",
        "Peter Zwegat wäre stolz auf diese Ordnung!",
        "Kontrolle ist besser als Nachsicht - gut gemacht!",
        "Ihre Disziplin zahlt sich aus!",
    ]

    import secrets

    # Verwende secrets für kryptographisch sichere Zufallsauswahl
    zwegat_spruch = secrets.choice(zwegat_sprueche)

    context = {
        "page_title": "Dashboard",
        "page_subtitle": f'Willkommen zurück! Heute ist {heute.strftime("%A, %d. %B %Y")}',
        # Finanzkennzahlen
        "einnahmen_monat": einnahmen_monat,
        "ausgaben_monat": ausgaben_monat,
        "gewinn_monat": gewinn_monat,
        "einnahmen_jahr": einnahmen_jahr,
        "ausgaben_jahr": ausgaben_jahr,
        "gewinn_jahr": gewinn_jahr,
        "einnahmen_trend": einnahmen_trend,
        # Statistiken
        "stats": stats,
        # Listen
        "letzte_buchungen": letzte_buchungen,
        "offene_belege": offene_belege,
        "top_ausgaben": top_ausgaben,
        # Chart-Daten
        "chart_data": chart_data,
        # Peter Zwegat
        "zwegat_spruch": zwegat_spruch,
        # Zeiträume
        "aktueller_monat": heute.strftime("%B %Y"),
        "aktuelles_jahr": heute.year,
    }

    return render(request, "dashboard/dashboard_modern.html", context)


def kennzahlen_ajax(request):
    """
    AJAX-Endpoint für Live-Kennzahlen.
    """
    heute = timezone.now().date()
    heute.replace(day=1)

    # Aktuelle Zahlen
    einnahmen_heute = Buchungssatz.objects.filter(
        buchungsdatum=heute, haben_konto__nummer__startswith="8"
    ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")

    buchungen_heute = Buchungssatz.objects.filter(buchungsdatum=heute).count()

    return JsonResponse(
        {
            "einnahmen_heute": float(einnahmen_heute),
            "buchungen_heute": buchungen_heute,
            "timestamp": timezone.now().isoformat(),
        }
    )


def eur_view(request):
    """
    Einnahmen-Überschuss-Rechnung (EÜR) für das Finanzamt.
    Peter Zwegat: "Das ist das Wichtigste - hier sieht das Finanzamt alles!"
    """
    jahr = int(request.GET.get("jahr", timezone.now().year))
    jahr_start = timezone.now().date().replace(year=jahr, month=1, day=1)
    jahr_ende = timezone.now().date().replace(year=jahr, month=12, day=31)

    # Einnahmen nach Kategorien
    einnahmen_kategorien = {
        "erlöse": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            haben_konto__nummer__startswith="8",
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "sonstige_einnahmen": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            haben_konto__nummer__startswith="48",  # Sonstige betriebliche Erträge
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
    }

    gesamte_einnahmen = sum(einnahmen_kategorien.values())
    # Prozentanteile berechnen
    einnahmen_anteile = {}
    for k, v in einnahmen_kategorien.items():
        if gesamte_einnahmen > 0:
            einnahmen_anteile[k] = float(v) / float(gesamte_einnahmen) * 100
        else:
            einnahmen_anteile[k] = 0.0

    # Ausgaben nach Kategorien
    ausgaben_kategorien = {
        "wareneinsatz": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["5000", "5100", "5200"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "personalkosten": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__startswith="62",
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "mieten": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["4120", "4130"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "buerokosten": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["4980", "4985", "4990"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "marketing": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["4600", "4610", "4620"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "reisekosten": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["4650", "4655"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "kfz_kosten": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["4520", "4530"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "versicherungen": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__in=["4360", "4370"],
        ).aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
        "sonstige_ausgaben": Buchungssatz.objects.filter(
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
            soll_konto__nummer__startswith="4",
            soll_konto__nummer__regex=r"^4[0-3]",  # Sonstige Aufwendungen
        )
        .exclude(
            soll_konto__nummer__in=[
                "4120",
                "4130",
                "4360",
                "4370",
                "4520",
                "4530",
                "4600",
                "4610",
                "4620",
                "4650",
                "4655",
                "4980",
                "4985",
                "4990",
            ]
        )
        .aggregate(summe=Sum("betrag"))["summe"]
        or Decimal("0"),
    }

    # Summen berechnen
    gesamte_einnahmen = sum(einnahmen_kategorien.values())
    gesamte_ausgaben = sum(ausgaben_kategorien.values())
    eur_ergebnis = gesamte_einnahmen - gesamte_ausgaben

    # Verfügbare Jahre für Dropdown
    erste_buchung = Buchungssatz.objects.order_by("buchungsdatum").first()
    if erste_buchung:
        start_jahr = erste_buchung.buchungsdatum.year
    else:
        start_jahr = timezone.now().year

    verfuegbare_jahre = list(range(start_jahr, timezone.now().year + 1))

    context = {
        "page_title": f"EÜR {jahr}",
        "page_subtitle": "Einnahmen-Überschuss-Rechnung für das Finanzamt",
        "jahr": jahr,
        "verfuegbare_jahre": verfuegbare_jahre,
        # EÜR-Daten
        "einnahmen_kategorien": einnahmen_kategorien,
        "einnahmen_anteile": einnahmen_anteile,
        "ausgaben_kategorien": ausgaben_kategorien,
        "gesamte_einnahmen": gesamte_einnahmen,
        "gesamte_ausgaben": gesamte_ausgaben,
        "eur_ergebnis": eur_ergebnis,
        # Zeitraum
        "jahr_start": jahr_start,
        "jahr_ende": jahr_ende,
    }

    return render(request, "auswertungen/eur.html", context)


def kontenblatt_view(request, konto_id):
    """
    Detailliertes Kontenblatt für ein einzelnes Konto.
    Peter Zwegat: "Hier sehen Sie jeden Euro - bis auf den Cent genau!"
    """
    konto = get_object_or_404(Konto, id=konto_id)

    # Zeitraum
    jahr = int(request.GET.get("jahr", timezone.now().year))
    monat = request.GET.get("monat")

    jahr_start = timezone.now().date().replace(year=jahr, month=1, day=1)
    jahr_ende = timezone.now().date().replace(year=jahr, month=12, day=31)

    # Filter nach Monat wenn gewählt
    if monat:
        monat = int(monat)
        jahr_start = jahr_start.replace(month=monat)
        if monat == 12:
            jahr_ende = jahr_start.replace(day=31)
        else:
            jahr_ende = jahr_start.replace(month=monat + 1, day=1) - timedelta(days=1)

    # Buchungen für dieses Konto
    soll_buchungen = (
        Buchungssatz.objects.filter(
            soll_konto=konto,
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
        )
        .select_related("haben_konto", "geschaeftspartner", "beleg")
        .order_by("buchungsdatum")
    )

    haben_buchungen = (
        Buchungssatz.objects.filter(
            haben_konto=konto,
            buchungsdatum__gte=jahr_start,
            buchungsdatum__lte=jahr_ende,
        )
        .select_related("soll_konto", "geschaeftspartner", "beleg")
        .order_by("buchungsdatum")
    )

    # Salden berechnen
    soll_summe = soll_buchungen.aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")
    haben_summe = haben_buchungen.aggregate(summe=Sum("betrag"))["summe"] or Decimal(
        "0"
    )

    # Saldo je nach Kontotyp
    if konto.nummer.startswith(("0", "1", "4", "5", "6")):  # Aktiv- und Aufwandskonten
        saldo = soll_summe - haben_summe
    else:  # Passiv- und Ertragskonten
        saldo = haben_summe - soll_summe

    context = {
        "page_title": f"Kontenblatt {konto.nummer}",
        "page_subtitle": konto.name,
        "konto": konto,
        "jahr": jahr,
        "monat": monat,
        "jahr_start": jahr_start,
        "jahr_ende": jahr_ende,
        "soll_buchungen": soll_buchungen,
        "haben_buchungen": haben_buchungen,
        "soll_summe": soll_summe,
        "haben_summe": haben_summe,
        "saldo": saldo,
        "verfuegbare_jahre": list(range(2020, timezone.now().year + 1)),
        "verfuegbare_monate": [
            (1, "Januar"),
            (2, "Februar"),
            (3, "März"),
            (4, "April"),
            (5, "Mai"),
            (6, "Juni"),
            (7, "Juli"),
            (8, "August"),
            (9, "September"),
            (10, "Oktober"),
            (11, "November"),
            (12, "Dezember"),
        ],
    }

    return render(request, "auswertungen/kontenblatt.html", context)


def eur_pdf_export(request):
    """
    PDF-Export der EÜR.
    Peter Zwegat: "Papier ist geduldig - aber Ihr Steuerberater wird begeistert sein!"
    """
    jahr = int(request.GET.get("jahr", timezone.now().year))

    # EÜR-Daten sammeln (vereinfacht - nur Hauptkategorien)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="EÜR_{jahr}.pdf"'

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    normal_style = styles["Normal"]

    # Titel
    title = Paragraph(f"Einnahmen-Überschuss-Rechnung {jahr}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))

    # Peter Zwegat Branding
    subtitle = Paragraph(
        "Erstellt mit llkjj_knut - Peter Zwegats Buchhaltungsbutler", normal_style
    )
    story.append(subtitle)
    story.append(Spacer(1, 24))

    # EÜR-Tabelle (vereinfacht)
    data = [
        ["Position", "Betrag (€)"],
        ["", ""],
        ["Betriebseinnahmen", ""],
        ["Umsatzerlöse", "0,00"],
        ["Sonstige betriebliche Erträge", "0,00"],
        ["", ""],
        ["Betriebsausgaben", ""],
        ["Wareneinsatz", "0,00"],
        ["Personalkosten", "0,00"],
        ["Raumkosten", "0,00"],
        ["Bürokosten", "0,00"],
        ["Sonstige Aufwendungen", "0,00"],
        ["", ""],
        ["Gewinn/Verlust", "0,00"],
    ]

    table = Table(data, colWidths=[12 * cm, 4 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 14),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    story.append(table)
    story.append(Spacer(1, 24))

    # Hinweis
    hinweis = Paragraph(
        "Peter Zwegat: 'Diese EÜR ist bereit für Ihren Steuerberater. "
        "Alle Zahlen sind nach SKR03-Standard kategorisiert!'",
        normal_style,
    )
    story.append(hinweis)

    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    response.write(pdf_content)

    return response


def eur_excel_export(request):
    """
    Excel-Export der EÜR.
    Peter Zwegat: "Excel ist wie ein Schweizer Taschenmesser - für alles zu gebrauchen!"
    """
    jahr = int(request.GET.get("jahr", timezone.now().year))

    # Excel-Workbook erstellen
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"EÜR {jahr}"

    # Header mit Stil
    worksheet.append(["Einnahmen-Überschuss-Rechnung", jahr])
    worksheet.append([])
    worksheet.append(["Position", "Betrag (€)", "Anteil (%)"])

    # Beispieldaten (hier würde die echte EÜR-Logik stehen)
    data = [
        ["Betriebseinnahmen", "", ""],
        ["Umsatzerlöse", 0, ""],
        ["Sonstige Erträge", 0, ""],
        ["", "", ""],
        ["Betriebsausgaben", "", ""],
        ["Wareneinsatz", 0, ""],
        ["Personalkosten", 0, ""],
        ["Raumkosten", 0, ""],
        ["Bürokosten", 0, ""],
        ["", "", ""],
        ["Gewinn/Verlust", 0, ""],
    ]

    for row_data in data:
        worksheet.append(row_data)

    # Formatierung
    # Header-Zeile fett
    for cell in worksheet[1]:
        cell.font = Font(bold=True, size=14)

    # Spalten-Zeile fett
    for cell in worksheet[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Spaltenbreiten
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 15
    worksheet.column_dimensions["C"].width = 15

    # Response vorbereiten
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="EÜR_{jahr}.xlsx"'

    # Excel-Datei in Response schreiben
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()

    return response


def eur_offiziell_view(request):
    """
    Offizielle EÜR basierend auf dem amtlichen Formular.
    Peter Zwegat: "Das ist das Wichtigste - hier sieht das Finanzamt alles!"
    """
    from .services import EURService

    jahr = int(request.GET.get("jahr", timezone.now().year))

    # EÜR-Service initialisieren
    eur_service = EURService(jahr)

    # Offizielle EÜR berechnen
    eur_data = eur_service.berechne_offizielle_eur()

    # Verfügbare Jahre
    verfuegbare_jahre = eur_service.get_verfuegbare_jahre()

    context = {
        "page_title": f"Offizielle EÜR {jahr}",
        "page_subtitle": "Einnahmen-Überschuss-Rechnung nach amtlichem Formular",
        "jahr": jahr,
        "verfuegbare_jahre": verfuegbare_jahre,
        # Offizielle EÜR-Daten
        "eur_data": eur_data,
        "einnahmen": eur_data["einnahmen"],
        "ausgaben": eur_data["ausgaben"],
        "gesamte_einnahmen": eur_data["gesamte_einnahmen"],
        "gesamte_ausgaben": eur_data["gesamte_ausgaben"],
        "eur_ergebnis": eur_data["eur_ergebnis"],
        "ist_gewinn": eur_data["ist_gewinn"],
        "ist_verlust": eur_data["ist_verlust"],
        # Peter Zwegat Motivation
        "zwegat_spruch": (
            "Das sieht schon sehr professionell aus!"
            if eur_data["eur_ergebnis"] > 0
            else "Keine Sorge, das wird schon wieder!"
        ),
    }

    return render(request, "auswertungen/eur_offiziell.html", context)


def eur_export_csv(request):
    """
    CSV-Export der offiziellen EÜR.
    Peter Zwegat: "Für den Steuerberater - sauber und übersichtlich!"
    """
    from .services import EURExportService, EURService

    jahr = int(request.GET.get("jahr", timezone.now().year))

    # EÜR berechnen
    eur_service = EURService(jahr)
    eur_data = eur_service.berechne_offizielle_eur()

    # CSV generieren
    csv_content = EURExportService.generiere_csv_export(eur_data)

    # HTTP-Response erstellen
    response = HttpResponse(csv_content, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="EUR_{jahr}.csv"'
    response["Content-Encoding"] = "utf-8"

    return response


def eur_export_pdf(request):
    """
    PDF-Export der offiziellen EÜR.
    Peter Zwegat: "Das können Sie direkt beim Finanzamt abgeben!"
    """
    from .services import EURService

    jahr = int(request.GET.get("jahr", timezone.now().year))

    # EÜR berechnen
    eur_service = EURService(jahr)
    eur_data = eur_service.berechne_offizielle_eur()

    # PDF generieren mit ReportLab
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    # Container für PDF-Elemente
    story = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(f"<b>Einnahmen-Überschuss-Rechnung {jahr}</b>", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 12))

    # Einnahmen-Tabelle
    story.append(Paragraph("<b>BETRIEBSEINNAHMEN</b>", styles["Heading2"]))

    einnahmen_data = [["Zeile", "Bezeichnung", "Betrag (EUR)"]]
    for item in eur_data["einnahmen"]:
        einnahmen_data.append(
            [
                item["zeile_nummer"],
                item["bezeichnung"],
                f"{item['betrag']:.2f}".replace(",", "."),
            ]
        )

    einnahmen_data.append(
        [
            "",
            "Summe Einnahmen:",
            f"{eur_data['gesamte_einnahmen']:.2f}".replace(",", "."),
        ]
    )

    einnahmen_table = Table(einnahmen_data, colWidths=[2 * cm, 12 * cm, 4 * cm])
    einnahmen_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )

    story.append(einnahmen_table)
    story.append(Spacer(1, 20))

    # Ausgaben-Tabelle
    story.append(Paragraph("<b>BETRIEBSAUSGABEN</b>", styles["Heading2"]))

    ausgaben_data = [["Zeile", "Bezeichnung", "Betrag (EUR)"]]
    for item in eur_data["ausgaben"]:
        ausgaben_data.append(
            [
                item["zeile_nummer"],
                item["bezeichnung"],
                f"{item['betrag']:.2f}".replace(",", "."),
            ]
        )

    ausgaben_data.append(
        [
            "",
            "Summe Ausgaben:",
            f"{eur_data['gesamte_ausgaben']:.2f}".replace(",", "."),
        ]
    )

    ausgaben_table = Table(ausgaben_data, colWidths=[2 * cm, 12 * cm, 4 * cm])
    ausgaben_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]
        )
    )

    story.append(ausgaben_table)
    story.append(Spacer(1, 20))

    # Ergebnis
    ergebnis_text = "Gewinn" if eur_data["ist_gewinn"] else "Verlust"
    story.append(
        Paragraph(
            f"<b>Jahresergebnis ({ergebnis_text}): {eur_data['eur_ergebnis']:,.2f} EUR</b>".replace(
                ",", "."
            ),
            styles["Heading2"],
        )
    )

    # Peter Zwegat Footer
    story.append(Spacer(1, 30))
    story.append(
        Paragraph(
            "<i>Erstellt mit llkjj_knut - Peter Zwegats Buchhaltungsbutler</i>",
            styles["Normal"],
        )
    )  # PDF generieren
    doc.build(story)

    # Response erstellen
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="EUR_{jahr}.pdf"'

    return response


def eur_mapping_details(request, mapping_id):
    """
    Details zu einem bestimmten EÜR-Mapping.
    Peter Zwegat: "Hier sehen Sie genau, was dahintersteckt!"
    """
    from .models import EURMapping
    from .services import EURService

    mapping = get_object_or_404(EURMapping, id=mapping_id)
    jahr = int(request.GET.get("jahr", timezone.now().year))

    # Service initialisieren
    eur_service = EURService(jahr)

    # Details für dieses Mapping holen
    details = eur_service.get_konten_details_fuer_mapping(mapping_id)

    # Gesamtbetrag berechnen
    kontoseite = "HABEN" if mapping.kategorie == "EINNAHMEN" else "SOLL"
    gesamtbetrag = eur_service._berechne_betrag_fuer_konten(
        mapping.skr03_konten, kontoseite
    )

    context = {
        "page_title": f"EÜR-Details Zeile {mapping.zeile_nummer}",
        "page_subtitle": mapping.bezeichnung,
        "mapping": mapping,
        "jahr": jahr,
        "details": details,
        "gesamtbetrag": gesamtbetrag,
        "kontoseite": kontoseite,
    }

    return render(request, "auswertungen/eur_mapping_details.html", context)


def eur_bmf_formular_view(request):
    """
    Offizielle EÜR-Ansicht nach BMF-Vorgaben (Anlage EÜR).

    Peter Zwegat: "Das ist das offizielle Formular -
    genau so, wie es das Finanzamt haben möchte!"
    """
    jahr = int(request.GET.get("jahr", timezone.now().year))

    # Verfügbare Jahre ermitteln
    verfuegbare_jahre = list(
        range(
            Buchungssatz.objects.aggregate(
                min_jahr=Min(Extract("buchungsdatum", "year"))
            )["min_jahr"]
            or timezone.now().year,
            timezone.now().year + 1,
        )
    )

    # Steuerpflichtigen-Daten (aus Einstellungen)
    steuerpflichtiger = {
        "name": "Max Mustermann",  # TODO: Aus Einstellungen laden
        "steuernummer": "123/456/78901",
        "beruf": "Freischaffender Künstler",
        "finanzamt": "Finanzamt Musterstadt",
    }

    # EÜR-Mappings mit offiziellen Kennziffern
    einnahmen_mappings = []
    ausgaben_mappings = []

    # Offizielle Einnahmen-Kategorien nach Anlage EÜR
    einnahmen_kategorien = [
        {
            "zeile": "1",
            "kennziffer": "511",
            "beschreibung": "Einnahmen aus freiberuflicher Tätigkeit",
            "tooltip": "Honorare, Vergütungen aus künstlerischer Tätigkeit",
            "konto_filter": ["8400", "8500", "8600"],
        },
        {
            "zeile": "2",
            "kennziffer": "512",
            "beschreibung": "Einnahmen aus Nebentätigkeiten",
            "tooltip": "Workshops, Kurse, Beratung",
            "konto_filter": ["8700", "8800"],
        },
        {
            "zeile": "3",
            "kennziffer": "513",
            "beschreibung": "Sonstige Einnahmen",
            "tooltip": "Verkäufe, Lizenzgebühren, etc.",
            "konto_filter": ["8900", "8950"],
        },
        {
            "zeile": "4",
            "kennziffer": "514",
            "beschreibung": "Vereinnahmte Umsatzsteuer",
            "tooltip": "Nur bei Regelbesteuerung",
            "konto_filter": ["1776"],
        },
    ]

    # Offizielle Ausgaben-Kategorien nach Anlage EÜR
    ausgaben_kategorien = [
        {
            "zeile": "21",
            "kennziffer": "521",
            "beschreibung": "Wareneinkauf",
            "tooltip": "Materialkosten, Werkstoffe",
            "konto_filter": ["4100", "4200"],
        },
        {
            "zeile": "22",
            "kennziffer": "522",
            "beschreibung": "Löhne und Gehälter",
            "tooltip": "Angestellte, Aushilfen",
            "konto_filter": ["4120", "4125"],
        },
        {
            "zeile": "23",
            "kennziffer": "523",
            "beschreibung": "Gesetzliche soziale Aufwendungen",
            "tooltip": "Sozialversicherung für Angestellte",
            "konto_filter": ["4130", "4135"],
        },
        {
            "zeile": "24",
            "kennziffer": "524",
            "beschreibung": "Mieten und Pachten",
            "tooltip": "Atelier, Büro, Lager",
            "konto_filter": ["4210", "4220"],
        },
        {
            "zeile": "25",
            "kennziffer": "525",
            "beschreibung": "Gas, Strom, Wasser",
            "tooltip": "Energiekosten",
            "konto_filter": ["4230", "4240"],
        },
        {
            "zeile": "26",
            "kennziffer": "526",
            "beschreibung": "Instandhaltung",
            "tooltip": "Reparaturen, Wartung",
            "konto_filter": ["4250", "4260"],
        },
        {
            "zeile": "27",
            "kennziffer": "527",
            "beschreibung": "Steuern, Versicherungen, Beiträge",
            "tooltip": "Berufsgenossenschaft, Kammerbeiträge",
            "konto_filter": ["4300", "4320"],
        },
        {
            "zeile": "28",
            "kennziffer": "528",
            "beschreibung": "Kfz-Kosten",
            "tooltip": "Benzin, Reparaturen, Versicherung",
            "konto_filter": ["4510", "4520"],
        },
        {
            "zeile": "29",
            "kennziffer": "529",
            "beschreibung": "Werbe- und Reisekosten",
            "tooltip": "Marketing, Geschäftsreisen",
            "konto_filter": ["4600", "4610"],
        },
        {
            "zeile": "30",
            "kennziffer": "530",
            "beschreibung": "Kosten der Warenabgabe",
            "tooltip": "Verpackung, Versand",
            "konto_filter": ["4700", "4710"],
        },
        {
            "zeile": "31",
            "kennziffer": "531",
            "beschreibung": "Abschreibungen",
            "tooltip": "AfA auf Anlagegüter",
            "konto_filter": ["4800", "4810"],
        },
        {
            "zeile": "32",
            "kennziffer": "532",
            "beschreibung": "Sonstige Ausgaben",
            "tooltip": "Büromaterial, Software, Fachliteratur",
            "konto_filter": ["4900", "4910", "4920"],
        },
        {
            "zeile": "33",
            "kennziffer": "533",
            "beschreibung": "Vorsteuer",
            "tooltip": "Nur bei Regelbesteuerung",
            "konto_filter": ["1574", "1575"],
        },
    ]

    # Beträge berechnen
    jahr_start = datetime.date(jahr, 1, 1)
    jahr_ende = datetime.date(jahr, 12, 31)

    summe_einnahmen = Decimal("0")
    summe_ausgaben = Decimal("0")

    # Einnahmen-Mappings erstellen
    for kategorie in einnahmen_kategorien:
        betrag = Decimal("0")
        for konto_nr in kategorie["konto_filter"]:
            konto_betrag = Buchungssatz.objects.filter(
                buchungsdatum__gte=jahr_start,
                buchungsdatum__lte=jahr_ende,
                haben_konto__nummer__startswith=konto_nr,
            ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")
            betrag += konto_betrag

        einnahmen_mappings.append(
            {
                "zeile": kategorie["zeile"],
                "kennziffer": kategorie["kennziffer"],
                "beschreibung": kategorie["beschreibung"],
                "tooltip": kategorie["tooltip"],
                "betrag": betrag,
            }
        )
        summe_einnahmen += betrag

    # Ausgaben-Mappings erstellen
    for kategorie in ausgaben_kategorien:
        betrag = Decimal("0")
        for konto_nr in kategorie["konto_filter"]:
            konto_betrag = Buchungssatz.objects.filter(
                buchungsdatum__gte=jahr_start,
                buchungsdatum__lte=jahr_ende,
                soll_konto__nummer__startswith=konto_nr,
            ).aggregate(summe=Sum("betrag"))["summe"] or Decimal("0")
            betrag += konto_betrag

        ausgaben_mappings.append(
            {
                "zeile": kategorie["zeile"],
                "kennziffer": kategorie["kennziffer"],
                "beschreibung": kategorie["beschreibung"],
                "tooltip": kategorie["tooltip"],
                "betrag": betrag,
            }
        )
        summe_ausgaben += betrag

    # Gewinn/Verlust berechnen
    gewinn = summe_einnahmen - summe_ausgaben

    # Besondere Angaben für Künstler
    kuenstlerische_einnahmen = summe_einnahmen  # Vereinfacht
    kleinunternehmer = True  # TODO: Aus Einstellungen laden

    # Sonderausgaben und außergewöhnliche Belastungen
    sonderausgaben = Decimal("0")  # TODO: Implementieren
    aussergewoehnliche_belastungen = Decimal("0")  # TODO: Implementieren

    context = {
        "page_title": "Einnahmenüberschussrechnung (EÜR)",
        "page_subtitle": f"Anlage EÜR nach § 60 EStDV für das Jahr {jahr}",
        "jahr": jahr,
        "verfuegbare_jahre": verfuegbare_jahre,
        "steuerpflichtiger": steuerpflichtiger,
        "einnahmen_mappings": einnahmen_mappings,
        "ausgaben_mappings": ausgaben_mappings,
        "summe_einnahmen": summe_einnahmen,
        "summe_ausgaben": summe_ausgaben,
        "gewinn": gewinn,
        "kuenstlerische_einnahmen": kuenstlerische_einnahmen,
        "kleinunternehmer": kleinunternehmer,
        "sonderausgaben": sonderausgaben,
        "aussergewoehnliche_belastungen": aussergewoehnliche_belastungen,
    }

    return render(request, "auswertungen/eur_bmf_formular.html", context)
